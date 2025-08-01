# -*- coding: utf-8 -*-
"""
General purpose utility functions for protograf
"""
# lib
import csv
import collections
import copy
from functools import lru_cache
from itertools import zip_longest
import jinja2
import logging
import io
import os
import pathlib
from pathlib import Path
import string
from string import ascii_uppercase, digits
import sys
from urllib.parse import urlparse

# third-party
from openpyxl import load_workbook
from pymupdf import Point as muPoint, Matrix, Font as muFont
from pymupdf.utils import getColor
import requests
import xlrd

# local
from protograf.utils.constants import (
    CACHE_DIRECTORY,
    COLOR_NAMES,
    DEFAULT_FONT,
    STANDARD_CARD_SIZES,
    PAPER_SIZES,
)
from protograf.utils.fonts import builtin_font, FontInterface
from protograf.utils.messaging import feedback
from protograf.utils.support import to_units
from protograf.utils.structures import (
    DirectionGroup,
    GlobalDocument,
    Point,
    ShapeProperties,
    TemplatingType,
)
from protograf import globals

log = logging.getLogger(__name__)
DEBUG = False
MIN_ATTRIBUTES = ("scheme", "netloc")
BUILTIN_FONTS = ["Times-Roman", "Courier", "Helvetica"]


__alpha_to_decimal = {letter: pos for pos, letter in enumerate(ascii_uppercase, 1)}
__powers = (1, 26, 676)


def script_path():
    """Get the path for a script being called from command line.

    Doc Test:

    >>> R = script_path()
    >>> 'utils' in R.parts
    True
    """
    fname = os.path.abspath(sys.argv[0])
    if fname:
        return pathlib.Path(fname).resolve().parent


def load_data(datasource=None, **kwargs):
    """
    Load data from a 'tabular' source (CSV, XLS) into a dict
    """
    dataset = {}
    log.debug("Load data from a 'tabular' source (CSV, XLS) %s", datasource)
    if datasource:
        filename, file_ext = os.path.splitext(datasource)
        if not filename:
            feedback("Unable to process a file without valid filename!", True)
        if file_ext and _lower(file_ext) == ".csv":
            headers = kwargs.get("headers", None)
            selected = kwargs.get("selected", None)
            dataset = open_csv(datasource, headers=headers, selected=selected)
        elif file_ext and _lower(file_ext) in [".xls", ".xlsx"]:
            headers = kwargs.get("headers", None)
            selected = kwargs.get("selected", None)
            sheet = kwargs.get("sheet", 0)
            sheetname = kwargs.get("sheetname", None)
            cells = kwargs.get("cells", None)
            dataset = open_excel(
                datasource,
                sheet=sheet,
                sheetname=sheetname,
                cells=cells,
                headers=headers,
            )
        else:
            feedback('Unable to process a file %s of type "%s"' % (filename, file_ext))
    return dataset


def load_googlesheet(sheet, **kwargs):
    """
    Load data from a Google Sheet into a dict
    """
    data_list = []
    spreadsheet_id = sheet
    api_key = kwargs.get("api_key", None)
    if not api_key:
        feedback('Cannot access a Google Sheet without an "api_key"', True)
    sheet_name = kwargs.get("name", None)
    if not sheet_name:
        feedback('Using default tab name of "Sheet1"', False, True)
        sheet_name = "Sheet1"
    log.debug("Load data from a Google Sheet %s", sheet)

    if sheet:
        url = f"https://sheets.googleapis.com/v4/spreadsheets/{spreadsheet_id}/values/{sheet_name}?alt=json&key={api_key}"
        try:
            response = requests.get(url)
            response.raise_for_status()  # raise exception for HTTP errors
            raw_data = response.json()
            data_dict = json_strings_to_numbers(raw_data)
        except requests.exceptions.RequestException as err:
            feedback(f"Unable to load Google Sheet: {err}")
            return []

    if data_dict:
        _data_list = data_dict.get("values")
        if _data_list:
            keys = _data_list[0]  # get keys/names from first sub-list
            dict_list = []
            for row in _data_list[1:]:
                _dict = {}
                for idx, key in enumerate(keys):
                    # handles the "bug" that Sheet does not return all of a row
                    # if the last cell is "empty"
                    _dict[key] = row[idx] if idx < len(row) else ""
                dict_list.append(_dict)
            return dict_list

    return data_list


def grouper(n, iterable, fillvalue=None):
    """Group and return sets

    See:
        http://stackoverflow.com/questions/2990121/~
        how-do-i-loop-through-a-python-list-by-twos

    Use:
        for item1, item2, item3 in grouper(3, 'ABCDEFG', 'x'):

    Doc Test:

    >>> list(grouper(3, 'ABCDEFG', 'x'))
    [('A', 'B', 'C'), ('D', 'E', 'F'), ('G', 'x', 'x')]
    """

    args = [iter(iterable)] * n
    return zip_longest(fillvalue=fillvalue, *args)


def boolean_join(items):
    """Create a result from a Boolean concatenation

    Doc Test:

    >>> items = [True, '+', False]
    >>> boolean_join(items)
    False
    >>> items = [True, '|', False]
    >>> boolean_join(items)
    True
    >>> items = [True, None]
    >>> boolean_join(items)
    True
    """
    if not items or len(items) == 0:
        return None
    expr = ""
    for item in items:
        if item == "&" or item == "and" or item == "+":
            expr += " and "
        elif item == "|" or item == "or":
            expr += " or "
        elif item is not None:
            expr += "%s" % item
        else:
            pass  # ignore nones
    try:
        result = eval(expr)
    except NameError:
        return None
    return result


def _lower(value) -> str | None:
    """Convert value into a lowercase string without any space around it

    Doc Test:

    >>> _lower(None)

    >>> _lower(1)
    '1'
    >>> _lower('a')
    'a'
    >>> _lower('AbA')
    'aba'
    >>> _lower( 'aB ')
    'ab'
    """
    if value is None:
        return None
    try:
        return str(value).lower().strip()
    except Exception:
        raise ValueError(f"Cannot convert {value} into a string!")


def as_int(
    value,
    label: str = None,
    maximum: int = None,
    minimum: int = None,
    allow_none: bool = False,
) -> int:
    """Convert a value to an int

    Args:

    - value (Any): the value to be converted to a float
    - label (str): assigned as part of the error message to ID the type of value
    - maximum (int): the upper allowed value for the conversion
    - lower (int): the lower allowed value for the conversion
    - allow_none (bool): if True, return None if value is None

    Doc Test:

    >>> as_int(value='3', label='N')
    3

    # below cannot be tested because of sys.exit() in feedback()
    # >>> as_int(value='3', label='N', minimum=4)
    # FEEDBACK:: z is
    # >>> as_int(value='3', label='N', maximum=2)
    # FEEDBACK:: z is
    # >>> as_int(value='z', label='N')
    # FEEDBACK:: The N value "z" is not a valid integer!
    # >>> as_int(value='3.1', label='N')
    # FEEDBACK:: The N value "3.1" is not a valid integer!
    """
    if value is None or value == "" and allow_none:
        return value
    _label = f"{label} value " if label else "value "
    try:
        the_value = int(value)
        if minimum and the_value < minimum:
            feedback(
                f'The {_label}"{value}" integer is less than the minimum of {minimum}!',
                True,
            )
        if maximum and the_value > maximum:
            feedback(
                f'The {_label}"{value}" integer is more than the maximum of {maximum}!',
                True,
            )
        return the_value
    except (ValueError, Exception):
        feedback(f'The {_label}"{value}" is not a valid integer!!', True)


def as_bool(value, label: str = None, allow_none: bool = True) -> bool:
    """Convert a value to a Boolean

    Args:

    - value (Any): the value to be converted to a float
    - label (str): assigned as part of the error message to ID the type of value
    - allow_none (bool): if True, return None if value is None

    Doc Test:

    >>> as_bool(value='3', label='N')
    False
    >>> as_bool(value='Y', label='Y')
    True
    """
    TRUES = ["yes", "ja", "oui", "si", "y", "ya", "yep", "yeah", "true", "t", "1"]
    if value is None and allow_none:
        return value
    _label = f" for {label}" if label else " of"
    result = str(value).lower() in TRUES
    return result


def as_float(
    value, label: str, maximum: float = None, minimum: float = None, stop: bool = True
) -> float:
    """Set a value to an float; or end program if an invalid value and stop is True

    Args:

    - value (Any): the value to be converted to a float
    - label (str): assigned as part of the error message to ID the type of value
    - maximum (float): the upper allowed value for the conversion
    - lower (float): the lower allowed value for the conversion
    - stop (bool): if True, halt program and display error message

    Doc Test:

    >>> as_float(value='3', label='N')
    3.0

    # below cannot be tested because of sys.exit() in feedback()
    # >>> as_float(value='3', label='N', minimum=4)
    # FEEDBACK:: z is
    # >>> as_float(value='3', label='N', maximum=2)
    # FEEDBACK:: z is
    # >>> as_float(value='z', label='N')
    # FEEDBACK:: z is not a valid N integer!
    # >>> as_float(value='3.1', label='N')
    # FEEDBACK:: The value "3.1" for N is not a valid integer!
    """
    _label = f" for {label}" if label else " "
    try:
        the_value = float(value)
        if minimum and the_value < minimum:
            feedback(
                f'The "{value}"{_label} float value is less than the minimum of {minimum}!',
                stop,
            )
        if maximum and the_value > maximum:
            feedback(
                f'The "{value}"{_label} float value is more than the maximum of {maximum}!',
                stop,
            )
        return the_value
    except (ValueError, Exception):
        if stop:
            feedback(f'The value "{value}"{label} is not a valid float number!', True)
        else:
            return None


def as_point(value) -> list | Point:
    """Convert one or more tuples to a Point or list of Points

    Doc Test:

    >>> as_point((1,2))
    Point(x=1, y=2)
    >>> as_point([(1,2), (3,4)])
    [Point(x=1, y=2), Point(x=3, y=4)]
    """
    if value is None:
        return None
    elif isinstance(value, tuple):
        return Point(value[0], value[1])
    elif isinstance(value, list):
        items = []
        for item in value:
            if isinstance(item, tuple):
                items.append(Point(item[0], item[1]))
            else:
                raise ValueError(f"Cannot convert {item} into a Point!")
        return items
    else:
        raise ValueError(f"Cannot convert {value} into a Point!")


def json_strings_to_numbers(json_data: str | dict | list):
    """Iteratively convert JSON data into numbers, if possible.

    Doc Test:

    >>> json_strings_to_numbers({})
    {}
    >>> json_strings_to_numbers([])
    []
    >>> json_strings_to_numbers('[]')
    '[]'
    >>> json_strings_to_numbers('["a", "1", "2.3"]')
    '["a", "1", "2.3"]'
    >>> json_strings_to_numbers(["a", "1", "2.3", {"a": "0.123"}])
    ['a', 1, 2.3, {'a': 0.123}]
    """
    if isinstance(json_data, dict):
        for key, value in json_data.items():
            json_data[key] = json_strings_to_numbers(value)
    elif isinstance(json_data, list):
        for i, item in enumerate(json_data):
            json_data[i] = json_strings_to_numbers(item)
    elif isinstance(json_data, str):
        try:
            return int(json_data)
        except ValueError:
            try:
                return float(json_data)
            except ValueError:
                return json_data
    return json_data


def tuple_split(
    string: str, label: str = "list", pairs_list: bool = False, all_ints: bool = False
) -> list:
    """Split a string into a list of tuple numbers

    Doc Test:

    >>> print(tuple_split(''))
    []
    >>> print(tuple_split('3'))
    [(3.0,)]
    >>> print(tuple_split('3,4'))
    [(3.0, 4.0)]
    >>> print(tuple_split('  3,4  5.1,6.2   -7,8.1'))
    [(3.0, 4.0), (5.1, 6.2), (-7.0, 8.1)]
    >>> print(tuple_split('3,5 6,1 4,2'))
    [(3.0, 5.0), (6.0, 1.0), (4.0, 2.0)]
    >>> print(tuple_split('3,5 6,1 4,2', all_ints=True))
    [(3, 5), (6, 1), (4, 2)]

    # below cannot be tested because of sys.exit() in feedback()
    # print(tuple_split('a,5 6,1 4,2', all_ints=True))
    # FEEDBACK:: Cannot convert list into a list of integer sets!
    # print(tuple_split('3,5 6,1 4', pairs_list=True))
    # Values of list must be pairs of integers!
    """
    values = []
    if string:
        try:
            _string_list = string.strip(" ").replace(";", ",").split(" ")
            # print(f'^^^ {_string_list=}')
            for _str in _string_list:
                items = _str.split(",")
                _items = []
                for itm in items:
                    _itm = itm.strip(" ")
                    if _itm:
                        if all_ints:
                            _items.append(int(_itm))
                        else:
                            _items.append(float(_itm))
                if _items:
                    values.append(tuple(_items))
            if pairs_list:
                for value in values:
                    if len(value) != 2:
                        feedback(
                            f"Values of {label} must be pairs of integers!",
                            f' Check if all values in "{string}" are integer pairs.',
                            True,
                        )
            return values
        except ValueError as err:
            if all_ints:
                feedback(
                    f"Cannot convert {label} into a list of integer sets!"
                    f' Check if all values in "{string}" are integers.',
                    True,
                )
            else:
                feedback(
                    f"Cannot convert {label} into a list of numeric sets ({err})!", True
                )
            return values

        except Exception:
            return values
    else:
        return values


def sequence_split(
    string: str,
    as_int: bool = True,
    unique: bool = True,
    sep: str = ",",
    as_float: bool = False,
    msg: str = "",
    clean: bool = False,
    star: bool = False,
) -> list:
    """
    Split a string into a list of individual values

    Args:

    - string: the item to be split
    - as_int (bool): if True, convert values to integers
    - unique (bool): if True, create a list of unique
    - sep (str): expected delimiter between values - defaults to ","
    - as_float (bool): if True, convert values to floats
    - msg (str): return as part of the error
    - clean (bool): if True, strip any surrounding spaces
    - star (bool): if True, allow for "all" or "*" as the only list value

    Note:
        * If `unique` is True, order will NOT be maintained!

    Doc Test:

    >>> sequence_split('*', star=True)
    ['*']
    >>> sequence_split('')
    []
    >>> sequence_split('3')
    [3]
    >>> sequence_split('3', as_int=False)
    ['3']
    >>> sequence_split('3,4,5')
    [3, 4, 5]
    >>> sequence_split('3,4,5', as_int=False, unique=False)
    ['3', '4', '5']
    >>> x = sequence_split('3,4,5', as_int=False)
    >>> assert '5' in x
    >>> sequence_split('3-5,6,1-4')
    [1, 2, 3, 4, 5, 6]
    >>> sequence_split('A,1,B', as_int=False, unique=False)
    ['A', '1', 'B']
    >>> sequence_split('3.1,4.2,5.3', unique=False, as_int=False, as_float=True)
    [3.1, 4.2, 5.3]
    >>> sequence_split([3.1,4.2,5.3], unique=False, as_int=False, as_float=True)
    [3.1, 4.2, 5.3]
    >>> sequence_split(3)
    [3]
    >>> sequence_split(3.1)
    [3.1]
    """
    values = []
    if isinstance(string, (dict, list)):
        return string
    if isinstance(string, (int, float)):
        return [string]
    if string:
        try:
            if sep == ",":
                _string = string.replace('"', "").replace("'", "").strip()
            else:
                _string = string
                if clean or star:
                    _string = _string.strip()
        except Exception:
            return values
    else:
        return values

    # simple single value
    try:
        if as_int:
            values.append(int(_string))
            return values
    except Exception:
        pass

    # multi-values
    try:
        _strings = _string.split(sep)
    except AttributeError as err:
        feedback(
            f'Unable to split "{_string}" - please check that its a valid candidate!',
            False,
        )
        if isinstance(_string, TemplatingType):
            feedback(f"The script may not be using T() correctly", True)
        else:
            feedback(f"", True)

    # star test
    if star and len(_strings) == 1:
        if _strings[0] == "*" or _strings[0] == "all":
            return ["*"]

    # log.debug('strings:%s', _strings)
    for item in _strings:
        if "-" in item:
            _strs = item.split("-")
            seq_range = [str(val) for val in _strs]
            if as_int:
                seq_range = list(range(int(_strs[0]), int(_strs[1]) + 1))
            values = values + seq_range
            if as_float:
                feedback(f'Cannot set a range of decimal numbers ("{item}"){msg}', True)
        else:
            if as_int:
                values.append(int(item))
            elif as_float:
                values.append(float(item))
            else:
                _item = str(item).strip() if clean else str(item)
                values.append(_item)

    if unique:
        return list(set(values))  # unique
    return values


def split(
    string: str, tuple_to_list: bool = False, separator: str = None, clean: bool = False
):
    """
    Split a string into a list of individual characters

    Doc Test:

    >>> split('A,1,B')
    ['A', '1', 'B']
    >>> split('A 1 B')
    ['A', '1', 'B']
    >>> split((1, 2, 3), True)
    [(1, 2, 3)]
    >>> split("1;2;3", separator=';')
    ['1', '2', '3']
    >>> split("1; 2; 3", separator=';', clean=True)
    ['1', '2', '3']
    >>> split("A,b B, C")
    ['A', 'b B', ' C']
    >>> split("A,b B, C", clean=True)
    ['A', 'b B', 'C']
    """
    if isinstance(string, list):
        return string
    if isinstance(string, tuple):
        if tuple_to_list:
            return [string]
        return string
    if separator:
        sep = separator
    else:
        sep = " " if string and "," not in string else ","
    return sequence_split(string, as_int=False, unique=False, sep=sep, clean=clean)


def integer_pairs(pairs, label: str = "list") -> list:
    """Convert a list or string into a list of tuples; each with a pair of integers.

    Doc Test:

    >>> integer_pairs(pairs=[(1,2), (3,4)])
    [(1, 2), (3, 4)]
    >>> integer_pairs(pairs="1,2 3,4")
    [(1, 2), (3, 4)]
    """
    if pairs:
        if isinstance(pairs, str):
            pairs = tuple_split(pairs, label=label, all_ints=True, pairs_list=True)
        if not isinstance(pairs, list):
            feedback(f"The {label} value '{pairs}' is not valid list!", True)
        for item in pairs:
            if not isinstance(item, tuple):
                feedback(
                    f'{label} must only contain a list of integers pairs (not "{pairs}")!',
                    True,
                )
            if len(item) != 2:
                feedback(
                    f'{label} must only contain a list of paired integers (not "{pairs}")!',
                    True,
                )
            for val in item:
                if not isinstance(val, int):
                    feedback(
                        f"{label} must only contain integers "
                        f' ("{val}" in "{pairs}" is not an integer)!',
                        True,
                    )
        return pairs
    return []


def splitq(seq, sep=None, pairs=("()", "[]", "{}"), quote="\"'"):
    """Split sequence by separator but considering parts inside pairs or quoted
       as unbreakable pairs have different start and end value, quote have same
       symbol in beginning and end.

    Notes:
        * Use itertools.islice if only part of splits is needed

    Source:
        https://www.daniweb.com/programming/software-development/code/426990/\
        split-string-except-inside-brackets-or-quotes

    Doc Test:

    >>> # TODO
    """
    if not seq:
        yield []
    else:
        lsep = len(sep) if sep is not None else 1
        lpair, _ = zip(*pairs)
        pairs = dict(pairs)
        start = index = 0
        while 0 <= index < len(seq):
            c = seq[index]
            if (sep and seq[index:].startswith(sep)) or (sep is None and c.isspace()):
                yield seq[start:index]
                # pass multiple separators as single one
                if sep is None:
                    index = len(seq) - len(seq[index:].lstrip())
                else:
                    while sep and seq[index:].startswith(sep):
                        index = index + lsep
                start = index
            elif c in quote:
                index += 1
                p, index = index, seq.find(c, index) + 1
                if not index:
                    raise IndexError("Unmatched quote %r\n%i:%s" % (c, p, seq[:p]))
            elif c in lpair:
                nesting = 1
                while True:
                    index += 1
                    p, index = index, seq.find(pairs[c], index)
                    if index < 0:
                        raise IndexError(
                            "Did not find end of pair for %r: %r\n%i:%s"
                            % (c, pairs[c], p, seq[:p])
                        )
                    nesting += "{lpair}({inner})".format(
                        lpair=c, inner=splitq(seq[p:index].count(c) - 2)
                    )
                    if not nesting:
                        break
            else:
                index += 1
        if seq[start:]:
            yield seq[start:]


def open_csv(filename: str = None, headers: list = None, selected: list = None):
    """Read data from CSV file into a list of dictionaries

    Args:

    - filename (str): path to CSV file
    - headers (list): a list of strings to use instead of the first row
    - selected (list): a list of desired rows e.g. [2,4,7]
    """
    if not filename:
        feedback("A valid CSV filename must be supplied!")

    dict_list = []
    _file_with_path = None
    norm_filename = os.path.normpath(filename)
    if not os.path.exists(norm_filename):
        filepath = script_path()
        _file_with_path = os.path.join(filepath, norm_filename)
        if not os.path.exists(_file_with_path):
            feedback(f'Unable to find CSV "{filename}", including in {filepath}')

    try:
        csv_filename = _file_with_path or norm_filename
        if headers:
            reader = csv.DictReader(open(csv_filename), fieldnames=headers)
        else:
            reader = csv.DictReader(open(csv_filename))
        for key, item in enumerate(reader):
            if not selected:
                dict_list.append(item)
            else:
                if key + 1 in selected:
                    dict_list.append(item)
    except IOError:
        feedback('Unable to find or open CSV "%s"' % csv_filename)
    return dict_list


def open_excel(
    filename: str,
    sheet: int = 0,
    sheetname: str = None,
    cells: str = None,
    headers: list = None,
):
    """Read data from an Excel file into a list of dictionaries

    Args:

    - filename (str): path to the file
    - sheet (int): select a sheet number (otherwise first is used)
    - sheetname (str): select a sheet by name (otherwise first is used)
    - cells (str): a range of cells delimiting data in the col:row format
      from top-left to bottom-right e.g. 'A3:E12'
    - headers (list): strings to use instead of the first row
    """

    def cleaned(value):
        if isinstance(value, float):
            if float(value) == float(int(value)):
                return int(value)
        result = "" if value is None else value
        return result

    if not filename:
        feedback("A valid Excel filename must be supplied!")

    _file_with_path = None
    norm_filename = os.path.normpath(filename)
    if not os.path.exists(norm_filename):
        filepath = script_path()
        _file_with_path = os.path.join(filepath, norm_filename)
        if not os.path.exists(_file_with_path):
            feedback(f'Unable to find "{filename}", including in {filepath}')
    else:
        _file_with_path = norm_filename  # in same dir as script!

    if sheet is None and sheetname is None:
        feedback(
            f'Access to Excel file "{filename}" requires either the sheet number or the sheet name!',
            True,
        )

    _, file_ext = os.path.splitext(_file_with_path)
    if file_ext == ".xls":
        return open_xls(
            filename=filename,
            sheet=sheet,
            sheetname=sheetname,
            cells=cells,
            headers=headers,
        )
    elif file_ext == ".xlsx":
        return open_xlsx(
            filename=filename,
            sheet=sheet,
            sheetname=sheetname,
            cells=cells,
            headers=headers,
        )
    else:
        feedback(f'Cannot process data files with an extension of "{file_ext}".', True)


def open_xlsx(
    filename: str,
    sheet: int = 0,
    sheetname: str = None,
    cells: str = None,
    headers: list = None,
):
    """Read data from XLSX file into a list of dictionaries

    Args:

    - filename (str): path to the file
    - sheet (int): select a sheet number (otherwise first is used)
    - sheetname (str): select a sheet by name (otherwise first is used)
    - headers (list): strings to use instead of the first row
    - cells (str): a range of cells delimiting data in the col:row format
      from top-left to bottom-right e.g. 'A3:E12'
    """

    def cleaned(value):
        if isinstance(value, float):
            if float(value) == float(int(value)):
                return int(value)
        result = "" if value is None else value
        return result

    if not filename:
        feedback("A valid Excel filename must be supplied!")

    dict_list = []

    _file_with_path = None
    norm_filename = os.path.normpath(filename)
    if not os.path.exists(norm_filename):
        filepath = script_path()
        _file_with_path = os.path.join(filepath, norm_filename)
        if not os.path.exists(_file_with_path):
            feedback(f'Unable to find "{filename}", including in {filepath}')

    try:
        excel_filename = _file_with_path or norm_filename
        book = load_workbook(excel_filename, data_only=True)
        if sheetname:
            sheet = book[sheetname]
        elif sheet is not None:
            sheet = int(sheet) - 1 if sheet > 0 else int(sheet)
            _sheetname = book.sheetnames[sheet]
            sheet = book[_sheetname]
        else:
            feedback(
                f'Access to Excel file "{filename}" requires either the sheet number or the sheet name!',
                True,
            )
        start = 1

        if not headers:
            keys = [
                sheet.cell(1, col_index).value
                for col_index in range(1, sheet.max_column + 1)
            ]
            if None in keys:
                feedback(
                    f'Please assign headers to all columns in Excel file "{filename}"'
                    " and/or delete any empty columns.",
                    True,
                )
            start = 2
        else:
            start = 1
            keys = headers
        if len(keys) < sheet.max_column:
            feedback(
                'Too few headers supplied for the existing columns in "%s"' % filename
            )
        else:
            dict_list = []
            if cells:
                _cells = cells.split(":")
                cell_range = sheet[_cells[0] : _cells[1]]
                for row_index, row in enumerate(cell_range):
                    item = {
                        keys[col_index]: cleaned(cell.value)
                        for col_index, cell in enumerate(row)
                    }
                    dict_list.append(item)
            else:
                for row_index in range(start, sheet.max_row + 1):
                    item = {
                        keys[col_index - 1]: cleaned(
                            sheet.cell(row=row_index, column=col_index).value
                        )
                        for col_index in range(1, sheet.max_column + 1)
                    }
                    dict_list.append(item)
    except IOError:
        feedback('Unable to find or open Excel "%s"' % excel_filename)
    except IndexError:
        feedback('Unable to open sheet "%s"' % (sheet or sheetname))
    except Exception:
        feedback('Unable to open or process sheet "%s"' % (sheet or sheetname))
    return dict_list


def open_xls(
    filename: str,
    sheet: int = 0,
    sheetname: str = None,
    cells: str = None,
    headers: list = None,
):
    """Read data from XLS file into a list of dictionaries

    Args:

    - filename (str): path to the file
    - sheet (int): select a sheet number (otherwise first is used)
    - sheetname (str): select a sheet by name (otherwise first is used)
    - headers (list): strings to use instead of the first row
    - cells (str): a range of cells delimiting data in the col:row format
      from top-left to bottom-right e.g. 'A3:E12'
    """

    def cleaned(value):
        if isinstance(value, float):
            if float(value) == float(int(value)):
                return int(value)
        # if isinstance(value, str):
        #     return value.encode('utf8')
        return value

    if not filename:
        feedback("A valid Excel filename must be supplied!")

    dict_list = []

    _file_with_path = None
    norm_filename = os.path.normpath(filename)
    if not os.path.exists(norm_filename):
        filepath = script_path()
        _file_with_path = os.path.join(filepath, norm_filename)
        if not os.path.exists(_file_with_path):
            feedback(f'Unable to find "{filename}", including in {filepath}')

    try:
        excel_filename = _file_with_path or norm_filename
        book = xlrd.open_workbook(excel_filename)
        if sheet:
            sheet = sheet - 1
            sheet = book.sheet_by_index(sheet)
        elif sheetname:
            sheet = book.sheet_by_name(sheetname)
        else:
            sheet = book.sheet_by_index(0)
        start = 1
        if not headers:
            keys = [sheet.cell(0, col_index).value for col_index in range(sheet.ncols)]
        else:
            start = 0
            keys = headers
        if len(keys) < sheet.ncols:
            feedback(
                'Too few headers supplied for the existing columns in "%s"' % filename
            )
        else:
            dict_list = []
            if cells:
                _cells = cells.split(":")
                _topleft = coordinate_to_tuple(_cells[0], True)  # (col, row)
                _btmrite = coordinate_to_tuple(_cells[1], True)  # (col, row)
                for row_index in range(_topleft[1], _btmrite[1] + 1):
                    item = {
                        keys[col_index]: cleaned(sheet.cell(row_index, col_index).value)
                        for col_index in range(_topleft[0], _btmrite[0] + 1)
                    }
                    dict_list.append(item)
            else:
                for row_index in range(start, sheet.nrows):
                    item = {
                        keys[col_index]: cleaned(sheet.cell(row_index, col_index).value)
                        for col_index in range(sheet.ncols)
                    }
                    dict_list.append(item)
    except IOError:
        feedback('Unable to find or open Excel "%s"' % excel_filename)
    except IndexError:
        feedback('Unable to open sheet "%s"' % (sheet or sheetname))
    except xlrd.biffh.XLRDError:
        feedback('Unable to open sheet "%s"' % sheetname)
    return dict_list


def flatten(lst: list):
    """Flatten nested lists into a single list

    Doc Test:

    >>> list(flatten([0, [1, 2], [3,4, [5,6]]]))
    [0, 1, 2, 3, 4, 5, 6]
    """
    try:
        for ele in lst:
            if isinstance(ele, collections.abc.Iterable) and not isinstance(ele, str):
                for sub in flatten(ele):
                    yield sub
            else:
                yield ele
    except TypeError:
        yield lst


def flatten_keys(d: dict):
    """Flatten nested dicts into a single dict.

    NOTE:
        * values for keys in nested dict will override those in parent(s)!
        * See: https://www.geeksforgeeks.org/python-flatten-nested-keys/

    Doc Test:

    >>> flatten_keys({'height': 8, 'cards': 1, 'image': None, 'kwargs': {'kwargs': {'image': 'FOO', 'kwargs': {'cards': 9}}}})
    {'height': 8, 'cards': 9, 'image': 'FOO'}
    """
    result = {}
    for k, v in d.items():
        if isinstance(v, dict):
            flat_v = flatten_keys(v)
            for flat_k, flat_v in flat_v.items():
                # result[k + '.' + flat_k] = flat_v
                result[flat_k] = flat_v
        elif isinstance(v, list):
            for i, item in enumerate(v):
                try:
                    flat_item = flatten_keys(item)
                except AttributeError:
                    flat_item = item
                if isinstance(flat_item, dict):
                    for flat_k, flat_v in flat_item.items():
                        # result[f"{k}.{i}.{flat_k}"] = flat_v
                        result[f"{flat_k}"] = flat_v
        else:
            result[k] = v
    return result


def comparer(val: str, operator: str, target: str | list) -> bool:
    """Compare value with a target.

    Args:

    - val (str): the value to be checked
    - operator (str): one of - < | > | ~ | *
    - target: a single value or a list of values; a list the operator must be a ~

    Doc Test:

    >>> comparer(None, None, None)
    True
    >>> comparer("1", '*', "1")
    FEEDBACK:: Unknown operator: * (1.0 and 1.0)
    False
    >>> comparer("1", None, "1")
    True
    >>> comparer("a", None, "a")
    True
    >>> comparer("True", None, "True")
    True
    >>> comparer("False", None, "False")
    True
    >>> comparer("1", '<', "1.1")
    True
    >>> comparer("a", '<', "aa")
    True
    >>> comparer("True", '<', "True")
    False
    >>> comparer("False", '<', "False")
    False
    >>> comparer("1", '~', "1.1")
    False
    >>> comparer("a", '~', "aa")
    True
    >>> comparer("True", '~', "True")
    False
    >>> comparer("False", '~', "False")
    False
    >>> comparer("1", '~', [1,2,3])
    True
    """

    def to_length(val, target):
        """Get length of object."""
        try:
            val = len(val)
        except Exception:
            pass
        try:
            target = len(target)
        except Exception:
            pass
        return val, target

    if target == "T" or target == "True":
        target = True
    if target == "F" or target == "False":
        target = False
    if val == "T" or val == "True":
        val = True
    if val == "F" or val == "False":
        val = False

    if not operator:
        operator = "="
    if operator in ["<", "<=", ">", ">="]:
        val, target = to_length(val, target)

    try:
        val = float(val)
    except Exception:
        pass
    try:
        target = float(target)
    except Exception:
        pass
    if operator == "=":
        if val == target:
            return True
    elif operator == "~" or operator == "in":
        try:
            if val in target:
                return True
        except TypeError:
            pass
    elif operator == "!=":
        if val != target:
            return True
    elif operator == "<":
        if val < target:
            return True
    elif operator == ">":
        if val > target:
            return True
    elif operator == ">=":
        if val >= target:
            return True
    elif operator == "<=":
        if val <= target:
            return True
    else:
        feedback("Unknown operator: %s (%s and %s)" % (operator, val, target))
    return False


def color_to_hex(name):
    """Convert a named color (Color class) to a hexadecimal string"""
    if isinstance(name, str):
        return name
    _tuple = (int(name.red * 255), int(name.green * 255), int(name.blue * 255))
    _string = "#%02x%02x%02x" % _tuple
    return _string.upper()


def rgb_to_hex(color: tuple) -> str:
    """Convert a RGB tuple color to a hexadecimal string

    Doc Test:

    >>> rgb_to_hex((123,45,6))
    '#7A852CD35FA'
    """
    if color is None:
        return color
    _tuple = (int(color[0] * 255), int(color[1] * 255), int(color[2] * 255))
    _string = "#%02x%02x%02x" % _tuple
    return _string.upper()


def alpha_column(num: int, lower: bool = False) -> string:
    """Convert a number to a letter-based notation

    Notes:
        * Encountered on a WarpWar map; numbers below 26 appear sequentially as
          a, b, c, etc, numbers above 26 appear sequentially as aa, bb, cc, etc; if
          above 52 then appear sequentially as aaa, bbb, ccc etc. Add more letters for
          each multiple of 26.

    Doc Test:

    >>> alpha_column(1)
    'A'
    >>> alpha_column(26, lower=True)
    'z'
    >>> alpha_column(27)
    'AA'
    """
    if lower:
        return string.ascii_lowercase[divmod(num - 1, 26)[1] % 26] * (
            divmod(num - 1, 26)[0] + 1
        )
    else:
        return string.ascii_uppercase[divmod(num - 1, 26)[1] % 26] * (
            divmod(num - 1, 26)[0] + 1
        )


@lru_cache(maxsize=None)
def column_from_string(col: str) -> int:
    """Convert ASCII column name (base 26) to decimal with 1-based index

    Characters represent descending multiples of powers of 26

    "AFZ" == 26 * pow(26, 0) + 6 * pow(26, 1) + 1 * pow(26, 2)

    Doc Test:

    >>> column_from_string('A')
    1
    >>> column_from_string('AA')
    27
    """
    error_msg = f"'{col}' is not a valid column name. Column names are from A to ZZZ"
    if len(col) > 3:
        raise ValueError(error_msg)
    idx = 0
    col = reversed(col.upper())
    for letter, power in zip(col, __powers):
        try:
            pos = __alpha_to_decimal[letter]
        except KeyError:
            raise ValueError(error_msg)
        idx += pos * power
    if not 0 < idx < 18279:
        raise ValueError(error_msg)
    return idx


def coordinate_to_tuple(coordinate: str, zeroed: bool = False) -> tuple:
    """Convert Excel style coordinate to 1-based (column, row) tuple

    Args:
        zeroed (bool): if True, use zero base

    Doc Test:

    >>> coordinate_to_tuple('A1')
    (1, 1)
    >>> coordinate_to_tuple('AB31')
    (28, 31)
    >>> coordinate_to_tuple('A1', True)
    (0, 0)
    >>> coordinate_to_tuple('AB31', True)
    (27, 30)
    """
    for idx, c in enumerate(coordinate):
        if c in digits:
            break
    col = coordinate[:idx]
    row = coordinate[idx:]
    if zeroed:
        return column_from_string(col) - 1, int(row) - 1
    else:
        return column_from_string(col), int(row)


def sheet_column(num: int, lower: bool = False) -> string:
    """Convert a spreadsheet number to a column letter

    Ref:
        https://stackoverflow.com/questions/23861680/

    Doc Test:

    >>> sheet_column(num=3, lower=True)
    'c'
    >>> sheet_column(num=27, lower=False)
    'AA'
    """

    def converter(num, lower):
        if lower:
            return (
                ""
                if num == 0
                else converter((num - 1) // 26, lower)
                + string.ascii_lowercase[(num - 1) % 26]
            )
        else:
            return (
                ""
                if num == 0
                else converter((num - 1) // 26, lower)
                + string.ascii_uppercase[(num - 1) % 26]
            )

    return converter(num, lower)


def get_font_by_name(font_name: str) -> tuple:
    """Get font details by name - built-in OR system installed.

    Args:
        font_name: expected name of font

    Returns:

    - font (pymupdf.Font): the Font object
    - font_file (str): path to the font's file
    - font_name (str): actual font name to be used
    - mu_font_name (str): font name to be used for HTML Text

    Doc Test:

    >>> get_font_by_name('foo')
    WARNING:: Cannot find or load a font named `foo`. Defaulting to "Helvetica".
    (Font('Helvetica'), None, 'Helvetica', 'Helvetica')
    >>> get_font_by_name('Helvetica')
    (Font('Helvetica'), None, 'Helvetica', 'Helvetica')

    #get_font_by_name('Arial')
    #(Font('Arial Regular'), '/usr/share/fonts/truetype/msttcorefonts/Arial.ttf', 'Arial')
    """

    font, font_file = None, None
    if not builtin_font(font_name):
        cache_directory = pathlib.Path(pathlib.Path.home() / CACHE_DIRECTORY)
        fi = FontInterface(cache_directory=cache_directory)
        font_file = fi.get_font_file(name=font_name)
        if not font_file:
            feedback(
                f"Cannot find or load a font named `{font_name}`."
                f' Defaulting to "{DEFAULT_FONT}".',
                False,
                True,
            )
            font_name = DEFAULT_FONT
            font = muFont(DEFAULT_FONT)  # built-in
        else:
            font = muFont(font_name, font_file)
    else:
        font = muFont(font_name)  # built-in

    mu_font_name = font_name.replace(" ", "-")
    return font, font_file, font_name, mu_font_name


def register_font(name: str, filename: str = None):
    """Register a font."""
    pass


def base_fonts():
    """Register MS Core Fonts

    NOTES:
        * On Ubuntu: sudo apt-get install ttf-mscorefonts-installer
        * The Windows filenames are 'truncated' versions, hence the use
          of an alternate
    """
    fonts = [
        {
            "name": "Arial",
            "alternate": "Arial",
        },
        {
            "name": "Verdana",
            "alternate": "Verda",
        },
        {
            "name": "Courier New",
            "alternate": "Cour",
        },
        {
            "name": "Times New Roman",
            "alternate": "Times",
        },
        {
            "name": "Trebuchet MS",
            "alternate": "Trebuc",
        },
        {
            "name": "Georgia",
            "alternate": "Georg",
        },
        {
            "name": "Webdings",
            "alternate": "Webd",
        },
        # {'name': 'ObiWan', 'alternate': 'benK'},  # dummy to check failure
    ]
    missing = []
    for ffont in fonts:
        try:
            name = ffont["name"]
            register_font(name)
        except Exception:
            try:
                alt = ffont.get("alternate")
                register_font(name, filename=alt)
            except Exception:
                missing.append(name)
    if missing:
        names = ", ".join(missing)
        feedback(f"Unable to register the MS font(s): {names}", False, True)


def eval_template(string: str, data: dict = None, label: str = ""):
    """Process data dict via jinja2 template in source.

    Doc Test:

    >>> eval_template("2+{{x}}", {'x': 2})
    '2+2'
    >>> eval_template("2+{{x}}", {'y': 2})
    '2+'
    """
    if data is None or not data:
        return string
    if isinstance(data, tuple):
        try:
            data = data._asdict()
        except Exception:
            pass
    if not isinstance(data, dict):
        feedback("The data must be in the form of a dictionary", True)
    try:
        environment = jinja2.Environment()
        template = environment.from_string(str(string))
        custom_value = template.render(data)
        return custom_value
    except jinja2.exceptions.TemplateSyntaxError:
        feedback(
            f'Unable to create the text or value - check the grammar for "{string}"',
            True,
        )
    except (ValueError, jinja2.exceptions.UndefinedError):
        feedback(f'Unable to process "{string}" data with this template', True)


def validated_directions(
    value: list | str, direction_group: DirectionGroup, label: str = ""
) -> list:
    """Check and return a list of lowercase, direction abbreviations.

    Doc Test:

    >>> validated_directions('n s', DirectionGroup.CARDINAL)
    ['n', 's']
    >>> validated_directions('ne se', DirectionGroup.HEX_FLAT)
    ['ne', 'se']
    >>> validated_directions('n s', DirectionGroup.HEX_POINTY)
    ['n', 's']
    >>> validated_directions('w e n s ne', DirectionGroup.COMPASS)
    ['w', 'e', 'n', 's', 'ne']
    >>> validated_directions(' w e n s ne ', DirectionGroup.COMPASS)  # spaces at ends
    ['w', 'e', 'n', 's', 'ne']
    """
    if not value:
        return []
    if isinstance(value, str):
        value = value.strip()
        values = split(value.lower())
    else:
        if not isinstance(value, list):
            feedback(f"Cannot handle {label}value - must be a string or a list!", True)
        values = [str(val).lower().strip() for val in value]
    values_set = set(values)
    match direction_group:
        case DirectionGroup.CARDINAL:
            valid = {"n", "e", "w", "s"}
        case DirectionGroup.COMPASS:
            valid = {"n", "e", "w", "s", "ne", "se", "sw", "nw"}
        case DirectionGroup.HEX_FLAT:
            valid = {"e", "se", "sw", "w", "ne", "nw"}
        case DirectionGroup.HEX_POINTY:
            valid = {"s", "se", "sw", "n", "ne", "nw"}
        case DirectionGroup.HEX_FLAT_EDGE:  # perbis
            valid = {"s", "se", "sw", "n", "ne", "nw"}
        case DirectionGroup.HEX_POINTY_EDGE:  # perbis
            valid = {"e", "se", "sw", "w", "ne", "nw"}
        case DirectionGroup.CIRCULAR:
            valid = {"n", "e", "w", "s", "ne", "se", "sw", "nw", "o", "d"}
        case _:
            raise NotImplementedError("Cannot handle {direction_group} type!")
    if "all" in values or "*" in values:
        values = list(valid)
        values_set = set(values)
    if values_set.issubset(valid):
        return values
    _label = f"the {label} value" if label else f'"{value}"'
    feedback(f"Cannot use {_label} - it must contain valid directions {valid}!", True)


def transpose_lists(
    original_list: list, direction: str = None, invert: str = None
) -> list:
    """Reorientate a list-of-lists

    Args:
        original_list (list)
            a list of lists
        direction (str)
            'south' / 's' / '-90' or  'north' / 'n' / '90' to swop rows with columns
        invert (str)
            'lr' or 'tb' to reverse order of sub-lists within outer list

    Doc Test:

    >>> transpose_lists([[1, 2, 3], [4, 5, 6]], direction=None, invert=None)
    [[1, 2, 3], [4, 5, 6]]
    >>> transpose_lists([[1, 2, 3], [4, 5, 6]], direction=None, invert='LR')
    [[3, 2, 1], [6, 5, 4]]
    >>> transpose_lists([[1, 2, 3], [4, 5, 6]], direction=None, invert='TB')
    [[4, 5, 6], [1, 2, 3]]
    >>> transpose_lists([[1, 2, 3], [4, 5, 6]], direction=90, invert=None)
    [[3, 6], [2, 5], [1, 4]]
    >>> transpose_lists([[1, 2, 3], [4, 5, 6]], direction=270, invert=None)
    [[4, 1], [5, 2], [6, 3]]
    >>> transpose_lists([[1, 2, 3], [4, 5, 6]], direction=90, invert='LR')
    [[1, 4], [2, 5], [3, 6]]
    >>> transpose_lists([[1, 2, 3], [4, 5, 6]], direction=270, invert='LR')
    [[6, 3], [5, 2], [4, 1]]
    >>> transpose_lists([[1, 2, 3], [4, 5, 6]], direction=90, invert='TB')
    [[6, 3], [5, 2], [4, 1]]
    >>> transpose_lists([[1, 2, 3], [4, 5, 6]], direction=270, invert='TB')
    [[1, 4], [2, 5], [3, 6]]
    >>> transpose_lists([[1,0], [1,0], [1,0], [1,1], ], direction=90, invert='TB')
    [[1, 0, 0, 0], [1, 1, 1, 1]]
    """

    def row_col_swop(matrix):
        num_cols = len(matrix[0])
        swopped_matrix = [[row[i] for row in matrix] for i in range(num_cols)]
        # print(f'^^^ {swopped_matrix=}')
        return swopped_matrix

    # print(f'^^^ transpose_lists {original_list=} {direction=} {invert=}')
    transpose_copy = copy.copy(original_list)
    match str(invert).lower():
        case "lr" | "leftright" | "rl" | "rightleft":
            for al in transpose_copy:
                al.reverse()
        case "tb" | "topbottom" | "bt" | "bottomtop":
            transpose_copy.reverse()
        case _:
            pass
    # print('^^^ PF post-flip', transpose_copy)

    match str(direction).lower():
        case "s" | "south" | "-90" | "270":
            transpose_copy = row_col_swop(transpose_copy)
            for al in transpose_copy:
                al.reverse()
        case "n" | "north" | "90":
            transpose_copy = row_col_swop(transpose_copy)
            transpose_copy.reverse()
        case _:
            pass
    # print('^^^ PF post-rotate', transpose_copy)
    return transpose_copy


def is_url_valid(url: str, qualifying=MIN_ATTRIBUTES):
    """Test if a URL is valid.

    See: https://stackoverflow.com/a/36283503/154858

    Doc Test:

    >>> is_url_valid(None)
    False
    >>> is_url_valid('')
    False
    >>> is_url_valid({})
    False
    >>> is_url_valid('naboo')
    False
    >>> is_url_valid('"file:///yoda.txt')
    False
    >>> is_url_valid('httpx://www.google.com')
    False
    >>> is_url_valid('https://https://https://www.foo.bar')
    False

    >>> is_url_valid('https://www.google.com')
    True
    >>> is_url_valid('https://www.tiktok.com/@outlikethevapors')
    True
    >>> is_url_valid('https://-wee.com')
    True
    >>> is_url_valid('http://localhost:8080')
    True
    """
    tokens = urlparse(url)
    if tokens.scheme and tokens.scheme not in ["http", "https"]:
        return False
    if tokens.netloc:
        if tokens.netloc[0:4] == "http" or tokens.netloc[0:5] == "https":
            return False
    return all(getattr(tokens, qualifying_attr) for qualifying_attr in qualifying)


def save_globals() -> GlobalDocument:
    """Create a copy of key globals settings"""
    return GlobalDocument(
        base=globals.base,
        deck=globals.deck,
        card_frames=globals.card_frames,
        filename=globals.filename,
        directory=globals.directory,
        document=globals.document,
        doc_page=globals.doc_page,
        canvas=globals.canvas,
        margins=globals.margins,
        page=globals.page,
        page_fill=globals.page_fill,
        page_width=globals.page_width,
        page_height=globals.page_height,
        page_count=globals.page_count,
        page_grid=globals.page_grid,
    )


def restore_globals(doc: GlobalDocument):
    """Restore key globals settings"""
    globals.base = doc.base
    globals.deck = doc.deck
    globals.card_frames = doc.card_frames
    globals.filename = doc.filename
    globals.directory = doc.directory
    globals.document = doc.document
    globals.doc_page = doc.doc_page
    globals.canvas = doc.canvas
    globals.margins = doc.margins
    globals.page = doc.page
    globals.page_fill = doc.page_fill
    globals.page_width = doc.page_width
    globals.page_height = doc.page_height
    globals.page_count = doc.page_count
    globals.page_grid = doc.page_grid


def get_color(name: str = None, is_rgb: bool = True) -> tuple:
    """Get a color tuple; by name from a pre-defined dictionary or as a RGB tuple."""
    if name is None:
        return None  # it IS valid to say that NO color has been set
    if isinstance(name, tuple) and len(name) == 3:  # RGB color tuple
        if (
            (name[0] >= 0 and name[0] <= 255)
            and (name[1] >= 0 and name[0] <= 255)
            and (name[2] >= 0 and name[0] <= 255)
        ):
            return name
        else:
            feedback(f'The color tuple "{name}" is invalid!')
    elif isinstance(name, str) and len(name) == 7 and name[0] == "#":  # hexadecimal
        _rgb = tuple(int(name[i : i + 2], 16) for i in (1, 3, 5))
        rgb = tuple(i / 255 for i in _rgb)
        return rgb
    else:
        pass  # unknown format
    try:
        if name.upper() not in COLOR_NAMES:
            feedback(f'The color name "{name}" is not pre-defined!', True)
        color = getColor(name)
        return color
    except (AttributeError, ValueError):
        feedback(f'The color name "{name}" cannot be converted to RGB!', True)


def get_opacity(transparency: float = 0) -> float:
    """Convert from '100% is fully transparent' to '0 is not opaque'."""
    if transparency is None:
        return 1.0
    try:
        return float(1.0 - transparency / 100.0)
    except (ValueError, TypeError):
        feedback(
            f'The transparency of "{transparency}" is not valid (use 0 to 100)', True
        )


def unit(item, units: str = None, skip_none: bool = False, label: str = ""):
    """Convert an item into the appropriate unit system."""
    log.debug("units %s :: label: %s", units, label)
    if item is None and skip_none:
        return None
    units = to_units(units) if units is not None else globals.units
    try:
        _item = as_float(item, label)
        return _item * units
    except (TypeError, ValueError):
        _label = f" {label}" if label else ""
        feedback(
            f"Unable to set unit value for{_label}: {item}."
            " Please check that this is a valid value.",
            stop=True,
        )


def get_pymupdf_props(
    defaults=None,
    index=None,  # extract from list of potential values (usually Card options)
    **kwargs,
):
    """Get pymupdf properties for fill, font, line, line style and colors

    Notes:
        If letting default a color parameter to None, then no resp. color selection
        command will be generated. If fill and color are both None, then the drawing
        will contain no color specification. But it will still be “stroked”,
        which causes PDF’s default color “black” be used by PDF viewers.

        The default value of width is 1.

        The values width, color and fill have the following relationship:

        - If fill=None, then shape elements will *always* be drawn with a border -
          even if color=None (in which case black is taken) or width=0
          (in which case 1 is taken).
        - Shapes without border can only be achieved if a fill color is specified
          (which may be be white). To achieve this, specify width=0.
          In this case, the color parameter is ignored.
    """

    def ext(prop):
        if isinstance(prop, str):
            return prop
        try:
            return prop[index]
        except TypeError:
            return prop

    defaults = defaults if defaults else {}
    # print(f'^^^ pymuProps: {kwargs.keys()} \n {kwargs.get("closed", "?")=}')
    if "fill" in kwargs.keys():
        fill = kwargs.get("fill", None)  # reserve None for 'no fill at all'
    else:
        fill = defaults.get("fill")
    if "stroke" in kwargs.keys():
        stroke = kwargs.get("stroke", None)  # reserve None for 'no stroke at all'
    else:
        stroke = defaults.get("stroke", None)
    # print(f'^^^ SCP {kwargs.get("fill")=} {fill=} {kwargs.get("stroke")=} {stroke=}')
    # ---- transparency / opacity
    opacity = 1
    _transparency = kwargs.get("transparency", defaults.get("transparency"))
    if _transparency:
        _transparency = as_float(_transparency, "transparency")
        if _transparency >= 1:
            _transparency = _transparency / 100.0
        opacity = 1 - _transparency
    stroke_width = kwargs.get("stroke_width", None)
    # stroke_cap = 1 if kwargs.get("stroke_cap", False) else 0  # rounded end line
    stroke_cap = 0
    if kwargs.get("rounded", False):
        stroke_cap = 1  # rounded end line
    if kwargs.get("squared", False):
        stroke_cap = 2  # semi-square; edge length of line width & center at line end
    dotted = kwargs.get("dotted", None)
    dashed = kwargs.get("dashed", None)
    _rotation = kwargs.get("rotation", None)  # calling Shape must set a tuple!
    _rotation_point = kwargs.get(
        "rotation_point", None
    )  # calling Shape must set a tuple!
    closed = kwargs.get("closed", False)  # whether to connect last and first points
    # ---- set line dots / dashed
    _dotted = ext(dotted) or ext(defaults.get("dotted"))
    _dashed = ext(dashed) or ext(defaults.get("dashed"))
    if _dotted:
        the_stwd = (
            round(ext(stroke_width))
            if stroke_width
            else round(ext(defaults.get("stroke_width")))
        )
        the_stwd = max(the_stwd, 1)
        dashes = f"[{the_stwd} {the_stwd}] 0"
    elif _dashed:
        _dlist = (
            _dashed
            if isinstance(_dashed, (list, tuple))
            else sequence_split(_dashed, as_int=False)
        )
        doffset = round(unit(_dlist[2])) if len(_dlist) >= 3 else 0
        dspaced = round(unit(_dlist[1])) if len(_dlist) >= 2 else ""
        dlength = round(unit(_dlist[0])) if len(_dlist) >= 1 else ""
        dashes = f"[{dlength} {dspaced}] {doffset}"
    else:
        dashes = None
    # print(f"### SCP{_dotted =} {_dashed=} {dashes=}")
    _line_join = kwargs.get("lineJoin", 0)  # set via rounded=True in PolylineShape
    # ---- check rotation
    morph = None
    # print(f'^^^ SCP {_rotation_point=} {_rotation}')
    if _rotation_point and not isinstance(_rotation_point, (Point, muPoint)):
        feedback(f'Rotation point "{_rotation_point}" is invalid', True)
    if _rotation is not None and not isinstance(_rotation, (float, int)):
        feedback(f'Rotation angle "{_rotation}" is invalid', True)
    if _rotation and _rotation_point:
        # ---- * set rotation matrix
        mtrx = Matrix(1, 1)
        mtrx.prerotate(_rotation)
        morph = (_rotation_point, mtrx)
        # print(f'^^^ SCP {morph=}')
    # ---- get color tuples
    _color = get_color(stroke)
    _fill = get_color(fill)
    # ---- set width
    _width = stroke_width or defaults.get("stroke_width")
    if _color is None and _fill is None:
        # feedback("Cannot have both fill and stroke set to None!", True)
        return None
    # print(f'^^^ SCP {stroke=} {fill=} {_color=} {_fill=}')  # None OR fraction RGB
    # ---- set/apply properties
    pymu_props = ShapeProperties(
        width=_width,
        color=_color,
        fill=_fill,
        lineCap=stroke_cap or 0,  # or self.stroke_cap,  # FIXME
        lineJoin=_line_join,
        dashes=dashes,
        fill_opacity=opacity,
        morph=morph,
        closePath=closed,
    )
    return pymu_props


def set_canvas_props(
    cnv=None,
    index=None,  # extract from list of potential values (usually Card options)
    defaults=None,
    **kwargs,
):
    """Set pymupdf Shape properties for fill, font, line, line style and colors"""
    # print(f'^^^ SetCnvProps: {kwargs.keys()} \n {kwargs.get("closed", "?")=}')
    cnv = cnv if cnv else globals.canvas
    defaults = defaults if defaults else {}
    pymu_props = get_pymupdf_props(defaults=defaults, index=index, **kwargs)
    if pymu_props:
        cnv.finish(
            width=pymu_props.width,
            color=pymu_props.color,
            fill=pymu_props.fill,
            lineCap=pymu_props.lineCap,
            lineJoin=pymu_props.lineJoin,
            dashes=pymu_props.dashes,
            fill_opacity=pymu_props.fill_opacity,
            morph=pymu_props.morph,
            closePath=pymu_props.closePath,
        )
    cnv.commit()
    return None


def get_font_file(font_name: str) -> tuple:
    """Access and track a font and its file."""
    _name = None
    font_path = None
    _file = None
    if not font_name:
        return _name, font_path, _file
    _font_name = str(font_name).strip()
    if _font_name:
        _name = builtin_font(_font_name)
        if not _name:  # check for custom font
            cache_directory = Path(Path.home() / CACHE_DIRECTORY)
            fi = FontInterface(cache_directory=cache_directory)
            _name = fi.get_font_family(font_name)
            if not _name:
                feedback(
                    f'Cannot find or load a font named "{font_name}".'
                    f' Defaulting to "{DEFAULT_FONT}".',
                    False,
                    True,
                )
            else:
                _file = fi.get_font_file(font_name, fullpath=False)
                font_path, css = fi.font_file_css(_name)
                if css not in globals.css:
                    globals.css += css + "\n"
                globals.archive.add(font_path)
                # print(font_path, globals.archive, globals.css)
    return _name, font_path, _file


def card_size(card_size: str = None, units: str = "pt") -> tuple:
    """Return card width and height in requested units for a named size.

    Doc Test:

    >>> card_size('poker')
    (180, 252)
    >>> card_size('poker', 'in')
    (2.5, 3.5)
    >>> card_size('miniamerican', 'mm')
    (41.0, 63.0)
    """
    size = None
    if units not in ["pt", "mm", "in"]:
        feedback(f'Card size units "{units}" is unknown.', True)
    match str(card_size).lower():
        case "bridge" | "b":
            size = STANDARD_CARD_SIZES["bridge"][units]
        case "business" | "u":
            size = STANDARD_CARD_SIZES["business"][units]
        case "mini" | "m":
            size = STANDARD_CARD_SIZES["mini"][units]
        case "miniamerican" | "ma":
            size = STANDARD_CARD_SIZES["miniamerican"][units]
        case "minieuropean" | "me":
            size = STANDARD_CARD_SIZES["minieuropean"][units]
        case "poker" | "p" | "mtg":
            size = STANDARD_CARD_SIZES["poker"][units]
        case "skat" | "s":
            size = STANDARD_CARD_SIZES["skat"][units]
        case "tarot" | "t":
            size = STANDARD_CARD_SIZES["tarot"][units]
        case "":
            pass
        case _:
            feedback(f'Card size "{card_size}" is unknown.', True)
    return size


def paper_size(paper_size: str = None, units: str = "pt") -> tuple:
    """Return paper width and height in requested units for a named size.

    Doc Test:

    >>> paper_size('A4')
    (595, 842)
    >>> paper_size('Legal', 'in')
    (8.5, 14)
    >>> paper_size('Notelet', 'pt')
    (270, 270)

    # >>> paper_size('A5', 'in')
    # (180, 252)
    """
    if units not in ["pt", "mm", "in"]:
        feedback(f'Paper size units "{units}" is unknown.', True)
    try:
        return PAPER_SIZES[paper_size][units]
    except KeyError:
        feedback(f'Paper size "{paper_size}" in "{units}" is unavailable.', True)


def uniques(key: str) -> list:
    """Unique values in a Data() dataset i.e. a list of equivalent dicts.

    Args:
        key (str): a key in each dict in the list of dicts
    """
    if not key:
        return []
    unique_values = set()
    for d in globals.dataset:
        if key in d:
            unique_values.add(d[key])
    return unique_values


if __name__ == "__main__":
    import doctest

    doctest.testmod()
